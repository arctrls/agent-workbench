#!/usr/bin/env python3
"""Convert a CSV file into a polished, compatibility-first XLSX workbook."""

from __future__ import annotations

import argparse
import csv
import io
import math
import re
import sys
import unicodedata
import zipfile
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError:
    print(
        "openpyxl is required. Install it in the selected Python environment "
        "with: python3 -m pip install openpyxl",
        file=sys.stderr,
    )
    raise SystemExit(2)


ENCODINGS = ("utf-8-sig", "utf-8", "cp949", "euc-kr", "utf-16")
DELIMITERS = ",\t;|"
TEXT_HEADER_MARKERS = (
    "id",
    "code",
    "sku",
    "barcode",
    "zip",
    "postal",
    "phone",
    "email",
    "코드",
    "바코드",
    "우편번호",
    "전화번호",
)
PERCENT_HEADER_MARKERS = ("%", "percent", "percentage", "rate", "ratio", "율")
DATE_HEADER_MARKERS = ("date", "day", "날짜", "일자")
DATETIME_HEADER_MARKERS = ("datetime", "timestamp", "_dt", "time", "일시", "시간")


@dataclass(frozen=True)
class ColumnSpec:
    kind: str
    number_format: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert one CSV file to a polished, Excel-compatible XLSX workbook."
    )
    parser.add_argument("input", type=Path, help="Input CSV path")
    parser.add_argument("--output", type=Path, help="Output XLSX path")
    parser.add_argument("--title", help="Workbook title shown above the data")
    parser.add_argument("--sheet-name", default="Data", help="Worksheet name")
    parser.add_argument(
        "--encoding",
        default="auto",
        help="CSV encoding or 'auto' (default: auto)",
    )
    parser.add_argument(
        "--delimiter",
        default="auto",
        help="CSV delimiter or 'auto' (default: auto)",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Overwrite an existing output file instead of adding a timestamp suffix",
    )
    return parser.parse_args()


def decode_csv(path: Path, requested_encoding: str) -> tuple[str, str]:
    raw = path.read_bytes()
    candidates = ENCODINGS if requested_encoding == "auto" else (requested_encoding,)
    errors: list[str] = []
    for encoding in candidates:
        try:
            return raw.decode(encoding), encoding
        except (LookupError, UnicodeDecodeError) as exc:
            errors.append(f"{encoding}: {exc}")
    raise ValueError("Unable to decode CSV. " + " | ".join(errors))


def detect_delimiter(text: str, requested_delimiter: str) -> str:
    if requested_delimiter != "auto":
        if requested_delimiter == r"\t":
            return "\t"
        if len(requested_delimiter) != 1:
            raise ValueError("--delimiter must be one character, '\\t', or 'auto'")
        return requested_delimiter
    sample = text[:65536]
    try:
        return csv.Sniffer().sniff(sample, delimiters=DELIMITERS).delimiter
    except csv.Error:
        return ","


def read_csv(
    path: Path, encoding: str, delimiter: str
) -> tuple[list[str], list[list[str]], str, str]:
    text, detected_encoding = decode_csv(path, encoding)
    detected_delimiter = detect_delimiter(text, delimiter)
    rows = list(csv.reader(io.StringIO(text), delimiter=detected_delimiter))
    while rows and not any(cell.strip() for cell in rows[-1]):
        rows.pop()
    if not rows:
        raise ValueError("CSV is empty")

    max_columns = max(len(row) for row in rows)
    normalized = [row + [""] * (max_columns - len(row)) for row in rows]
    headers = normalized[0]
    for index in range(len(headers)):
        if not headers[index].strip():
            headers[index] = f"Column {index + 1}"
    return headers, normalized[1:], detected_encoding, detected_delimiter


def normalize_header(header: str) -> str:
    return re.sub(r"[\s_-]+", "", header.strip().lower())


def looks_like_text_identifier(header: str, values: list[str]) -> bool:
    normalized_header = normalize_header(header)
    if any(normalized_header.endswith(marker) for marker in TEXT_HEADER_MARKERS):
        return True
    return any(re.fullmatch(r"[+-]?0\d+", value.strip()) for value in values if value.strip())


def parse_number(value: str) -> float | None:
    normalized = value.strip().replace(",", "").replace(" ", "")
    normalized = normalized.replace("₩", "").replace("$", "").replace("€", "")
    normalized = normalized.removesuffix("%")
    if not normalized:
        return None
    try:
        number = float(normalized)
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def parse_datetime_value(value: str) -> datetime | None:
    normalized = value.strip().replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        return None
    if parsed.tzinfo is not None:
        parsed = parsed.replace(tzinfo=None)
    return parsed


def parse_date_value(value: str) -> date | None:
    normalized = value.strip().replace("/", "-")
    try:
        return date.fromisoformat(normalized)
    except ValueError:
        return None


def infer_column(header: str, values: list[str]) -> ColumnSpec:
    nonempty = [value for value in values if value.strip()]
    if not nonempty or looks_like_text_identifier(header, nonempty):
        return ColumnSpec("text", "@")

    normalized_header = normalize_header(header)
    lowered = {value.strip().lower() for value in nonempty}
    if lowered <= {"y", "n", "yes", "no"}:
        return ColumnSpec("flag", "@")
    if any(marker in normalized_header for marker in DATETIME_HEADER_MARKERS):
        if all(parse_datetime_value(value) is not None for value in nonempty):
            return ColumnSpec("datetime", "yyyy-mm-dd hh:mm:ss")
    if any(marker in normalized_header for marker in DATE_HEADER_MARKERS):
        if all(parse_date_value(value) is not None for value in nonempty):
            return ColumnSpec("date", "yyyy-mm-dd")

    numbers = [parse_number(value) for value in nonempty]
    if all(number is not None for number in numbers):
        numeric_values = [number for number in numbers if number is not None]
        if any(normalized_header.endswith(marker) for marker in PERCENT_HEADER_MARKERS):
            decimal_places = 1 if any(not number.is_integer() for number in numeric_values) else 0
            return ColumnSpec("percent", "0.0%" if decimal_places else "0%")
        if all(number.is_integer() for number in numeric_values):
            return ColumnSpec("integer", "#,##0")
        return ColumnSpec("number", "#,##0.00")

    if lowered <= {"true", "false"}:
        return ColumnSpec("boolean", "General")
    return ColumnSpec("text", "@")


def convert_value(raw: str, spec: ColumnSpec) -> object:
    value = raw.strip()
    if not value:
        return None
    if spec.kind == "text":
        return raw
    if spec.kind == "integer":
        number = parse_number(value)
        return int(number) if number is not None else raw
    if spec.kind == "number":
        number = parse_number(value)
        return number if number is not None else raw
    if spec.kind == "percent":
        number = parse_number(value)
        if number is None:
            return raw
        return number / 100 if abs(number) > 1 else number
    if spec.kind == "date":
        return parse_date_value(value) or raw
    if spec.kind == "datetime":
        return parse_datetime_value(value) or raw
    if spec.kind == "boolean":
        return value.lower() == "true"
    return raw


def display_width(value: object) -> int:
    text = "" if value is None else str(value)
    return sum(2 if unicodedata.east_asian_width(char) in {"W", "F"} else 1 for char in text)


def safe_sheet_name(name: str) -> str:
    sanitized = re.sub(r"[\\/*?:\[\]]", "_", name).strip() or "Data"
    return sanitized[:31]


def choose_output_path(input_path: Path, requested: Path | None, overwrite: bool) -> Path:
    output = requested.expanduser() if requested else input_path.with_suffix(".xlsx")
    output = output.resolve()
    if output.suffix.lower() != ".xlsx":
        raise ValueError("Output path must end with .xlsx")
    if output.exists() and not overwrite:
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        output = output.with_name(f"{output.stem}-{timestamp}{output.suffix}")
    return output


def build_workbook(
    headers: list[str],
    raw_rows: list[list[str]],
    title: str,
    sheet_name: str,
    source_name: str,
) -> tuple[Workbook, list[ColumnSpec], list[list[object]]]:
    specs = [
        infer_column(header, [row[index] for row in raw_rows])
        for index, header in enumerate(headers)
    ]
    converted_rows = [
        [convert_value(value, specs[index]) for index, value in enumerate(row)]
        for row in raw_rows
    ]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = safe_sheet_name(sheet_name)
    worksheet.sheet_view.showGridLines = False
    worksheet.sheet_view.zoomScale = 90
    worksheet.freeze_panes = "A5"
    worksheet.sheet_properties.tabColor = "0F766E"

    workbook.properties.creator = "Codex"
    workbook.properties.title = title
    workbook.properties.subject = f"Converted from {source_name}"

    last_column = get_column_letter(len(headers))
    if len(headers) > 1:
        worksheet.merge_cells(f"A1:{last_column}1")
        worksheet.merge_cells(f"A2:{last_column}2")
    worksheet["A1"] = title
    worksheet["A2"] = f"{len(converted_rows):,} rows · {len(headers):,} columns · {source_name}"

    navy = "17324D"
    teal = "0F766E"
    white = "FFFFFF"
    slate = "425466"
    band = "E8F0F7"
    border_color = "D7DEE5"

    for cell_ref in ("A1", "A2"):
        worksheet[cell_ref].fill = PatternFill("solid", fgColor=navy)
        worksheet[cell_ref].alignment = Alignment(horizontal="left", vertical="center")
    worksheet["A1"].font = Font(name="Arial", size=18, bold=True, color=white)
    worksheet["A2"].font = Font(name="Arial", size=10, color="D7E5EF")

    header_row = 4
    data_start_row = 5
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=header_row, column=column_index, value=header)
        cell.fill = PatternFill("solid", fgColor=teal)
        cell.font = Font(name="Arial", size=10, bold=True, color=white)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style="medium", color="0A5B51"))

    thin_border = Side(style="thin", color=border_color)
    for row_offset, values in enumerate(converted_rows):
        row_number = data_start_row + row_offset
        for column_index, value in enumerate(values, start=1):
            cell = worksheet.cell(row=row_number, column=column_index, value=value)
            cell.font = Font(name="Arial", size=10, color=slate)
            cell.border = Border(bottom=thin_border)
            cell.number_format = specs[column_index - 1].number_format
            if row_offset % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=band)
            alignment = "left" if specs[column_index - 1].kind == "text" else "right"
            if specs[column_index - 1].kind in {"boolean", "flag"}:
                alignment = "center"
            cell.alignment = Alignment(horizontal=alignment, vertical="center")

    last_data_row = max(header_row, data_start_row + len(converted_rows) - 1)
    worksheet.auto_filter.ref = f"A{header_row}:{last_column}{last_data_row}"

    for column_index, header in enumerate(headers, start=1):
        values = [header] + [row[column_index - 1] for row in converted_rows[:500]]
        width = min(40, max(8, max(display_width(value) for value in values) + 2))
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.row_dimensions[1].height = 34
    worksheet.row_dimensions[2].height = 22
    worksheet.row_dimensions[3].height = 9
    worksheet.row_dimensions[4].height = 27
    for row_number in range(data_start_row, last_data_row + 1):
        worksheet.row_dimensions[row_number].height = 21

    worksheet.print_area = f"A1:{last_column}{last_data_row}"
    worksheet.print_title_rows = f"{header_row}:{header_row}"
    worksheet.page_setup.orientation = "landscape"
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_margins.left = 0.25
    worksheet.page_margins.right = 0.25
    worksheet.page_margins.top = 0.4
    worksheet.page_margins.bottom = 0.4

    return workbook, specs, converted_rows


def equivalent(actual: object, expected: object) -> bool:
    if isinstance(expected, date) and not isinstance(expected, datetime):
        if isinstance(actual, datetime):
            return actual.date() == expected
    if isinstance(expected, float):
        return isinstance(actual, (int, float)) and math.isclose(
            float(actual), expected, rel_tol=1e-12, abs_tol=1e-12
        )
    return actual == expected


def verify_output(
    output: Path,
    sheet_name: str,
    headers: list[str],
    converted_rows: list[list[object]],
) -> None:
    with zipfile.ZipFile(output) as archive:
        bad_member = archive.testzip()
        if bad_member:
            raise ValueError(f"Corrupt XLSX member: {bad_member}")

    workbook = load_workbook(output, data_only=False)
    worksheet = workbook[safe_sheet_name(sheet_name)]
    if worksheet.max_row != len(converted_rows) + 4 or worksheet.max_column != len(headers):
        raise ValueError("Reopened workbook dimensions do not match the CSV")
    if worksheet.tables or len(worksheet.conditional_formatting):
        raise ValueError("Compatibility-first workbook unexpectedly contains advanced XML features")
    for column_index, header in enumerate(headers, start=1):
        if worksheet.cell(row=4, column=column_index).value != header:
            raise ValueError(f"Header mismatch at column {column_index}")
    for row_index, expected_row in enumerate(converted_rows, start=5):
        for column_index, expected in enumerate(expected_row, start=1):
            actual = worksheet.cell(row=row_index, column=column_index).value
            if not equivalent(actual, expected):
                raise ValueError(
                    f"Value mismatch at row {row_index}, column {column_index}: "
                    f"{actual!r} != {expected!r}"
                )
    workbook.close()


def main() -> int:
    args = parse_args()
    input_path = args.input.expanduser().resolve()
    if not input_path.is_file():
        raise FileNotFoundError(f"CSV not found: {input_path}")
    if input_path.suffix.lower() != ".csv":
        raise ValueError("Input path must end with .csv")

    output_path = choose_output_path(input_path, args.output, args.overwrite)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    headers, raw_rows, encoding, delimiter = read_csv(
        input_path, args.encoding, args.delimiter
    )
    title = args.title or re.sub(r"[_-]+", " ", input_path.stem).strip()
    workbook, _, converted_rows = build_workbook(
        headers, raw_rows, title, args.sheet_name, input_path.name
    )
    workbook.save(output_path)
    verify_output(output_path, args.sheet_name, headers, converted_rows)

    printable_delimiter = "\\t" if delimiter == "\t" else delimiter
    print(f"Created: {output_path}")
    print(f"Rows: {len(converted_rows)}")
    print(f"Columns: {len(headers)}")
    print(f"Encoding: {encoding}")
    print(f"Delimiter: {printable_delimiter}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except (OSError, ValueError) as exc:
        print(f"error: {exc}", file=sys.stderr)
        raise SystemExit(1)
